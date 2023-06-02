import os
import re
import sys
import glob
import shutil
import datetime
import openpyxl
# from openpyxl.styles import PatternFill
from openpyxl.workbook.protection import WorkbookProtection
from tqdm import tqdm
import colorama
from colorama import Fore, Style


def printColor(color, string):
    fColor = ''
    if color == 'blue':
        fColor = Fore.CYAN
    elif color == 'red':
        fColor = Fore.RED
    elif color == 'green':
        fColor = Fore.GREEN
    elif color == 'yellow':
        fColor = Fore.YELLOW
    elif color == 'purple':
        fColor = Fore.MAGENTA 
        
    print(fColor + string + Style.RESET_ALL)


# path_of_factory_records = '../../../李/Factory Records/'
path_of_consumption_reports = './consumption reports'
path_of_factory_records = './'
hexiao_file = './auto核销.xlsx'
# password = '2400LI'


def main():
    colorama.init()
    pull = False
    add = False
    scrap = False

    pull = True
    add = True
    scrap = True

    consumption_reports = get_consumption_files(path_of_consumption_reports, 'Consumption Report')
    scrap_reports = get_consumption_files(path_of_consumption_reports, 'Scrap Report')

    for report in consumption_reports:
        process_consumption(report, add, pull)

    if scrap:
        for report in scrap_reports:
            process_scarp(report)

    print('complete')
    print("Press any key to continue...")
    input()


def get_consumption_files(path, report_type):
    matching_files = []

    # Iterate over all the files in the given path
    for file in os.listdir(path):
        # Check if the file is an xlsx file and its name starts with the report_type
        if file.endswith(".xlsx") and file.startswith(report_type):
            # If so, append the file name to the matching_files list
            file_path = os.path.join(path, file)
            matching_files.append(file_path)

    # Return the list of matching file names
    return matching_files

# def process_consumption_and_scarp(consumption_report, scrap_report):
#     # consumption_report = get_latest_consumption_report()
#     print(consumption_report)
    
#     # scrap_report = get_latest_scrap_report()
#     print(scrap_report)

#     consumpt_date = get_report_date(consumption_report, "WOConsumption")
#     scrap_date = get_report_date(scrap_report, "ConsumptionReport")

#     if consumpt_date == scrap_date:
#         print(scrap_date)
#     else:
#         print('reading different date')
#         print(consumpt_date)
#         print(scrap_date)
    
#     inventory_file = get_inventory_file(path_of_factory_records)
#     print(inventory_file)
    
#     backup_file(inventory_file)
#     backup_file(hexiao_file)


#     consumption = get_consumed_materials(consumption_report)
#     consumption = [material for material in consumption if material.get("material_number") != "B00"]
#     for i in range(0, len(consumption)):
#         if consumption[i]["material_number"] == "A36-COHO":
#             consumption[i]["material_number"] = "A36"
#             break
#     # print(consumption)

#     scraps = get_scraped_materials(scrap_report)

#     raw = [d for d in scraps if d['material_number'].startswith('A')]
#     add_scrap(inventory_file, "RAW OUT", raw)

#     ingredient = [d for d in scraps if d['material_number'].startswith('B')]
#     add_scrap(inventory_file, "INGREDIENT OUT", ingredient)

#     bag = [d for d in scraps if d['material_number'].startswith('C')]
#     add_scrap(inventory_file, "BAG OUT", bag)

#     box = [d for d in scraps if d['material_number'].startswith('D') or d['material_number'].startswith('E')]
#     add_scrap(inventory_file, "BOX OUT", box)

#     print('Getting pulled materials')
#     pulled_materials = get_pulled_materials(consumption_report)

#     print('Adding pulled materials')
#     add_pulled_materials(pulled_materials, hexiao_file)

#     print('Adding consumption')

#     raw = [d for d in consumption if d['material_number'].startswith('A')]
#     add_consumption(inventory_file, "RAW OUT", consumpt_date, raw)

#     ingredient = [d for d in consumption if d['material_number'].startswith('B')]
#     add_consumption(inventory_file, "INGREDIENT OUT", consumpt_date, ingredient)

#     bag = [d for d in consumption if d['material_number'].startswith('C')]
#     add_consumption(inventory_file, "BAG OUT", consumpt_date, bag)

#     box = [d for d in consumption if d['material_number'].startswith('D') or d['material_number'].startswith('E')]
#     add_consumption(inventory_file, "BOX OUT", consumpt_date, box)


def process_consumption(consumption_report, add = False, pull = False):
    print(consumption_report)

    consumpt_date = get_report_date(consumption_report, "WOConsumption")
    
    inventory_file = get_inventory_file(path_of_factory_records)
    # print(inventory_file)

    consumption = get_consumed_materials(consumption_report)
    consumption = [material for material in consumption if material.get("material_number") != "B00"]
    for i in range(0, len(consumption)):
        if consumption[i]["material_number"] == "A36-COHO":
            consumption[i]["material_number"] = "A36"
        if consumption[i]["material_number"] == "A36-KETA":
            consumption[i]["material_number"] = "A36"
        if consumption[i]["material_number"] == "B35":
            consumption[i]["material_number"] = "B22"
        if consumption[i]["material_number"] == "B19-F":
            consumption[i]["material_number"] = "B19"

    consumption = unifyMaterialNumbers(consumption, 'c')

    if pull:
        pulled_materials = get_pulled_materials(consumption_report)
        # add_pulled_materials_to_all(pulled_materials, hexiao_file)
        add_pulled_materials(pulled_materials, hexiao_file, consumpt_date)

    if add:
        raw = [d for d in consumption if d['material_number'].startswith('A')]
        add_consumption(inventory_file, "RAW OUT", consumpt_date, raw)

        ingredient = [d for d in consumption if d['material_number'].startswith('B')]
        add_consumption(inventory_file, "INGREDIENT OUT", consumpt_date, ingredient)

        bag = [d for d in consumption if d['material_number'].startswith('C')]
        add_consumption(inventory_file, "BAG OUT", consumpt_date, bag)

        box = [d for d in consumption if d['material_number'].startswith('D') or d['material_number'].startswith('E')]
        add_consumption(inventory_file, "BOX OUT", consumpt_date, box)


def process_scarp(scrap_report):
    print(scrap_report)
    scrap_date = get_report_date(scrap_report, "ConsumptionReport")
    
    inventory_file = get_inventory_file(path_of_factory_records)
    # print(inventory_file)

    scraps = get_scraped_materials(scrap_report)
    
    scraps = [material for material in scraps if material.get("material_number") != "B00"]
    for i in range(0, len(scraps)):
        if scraps[i]["material_number"] == "A36-COHO":
            scraps[i]["material_number"] = "A36"
        if scraps[i]["material_number"] == "A36-KETA":
            scraps[i]["material_number"] = "A36"
        if scraps[i]["material_number"] == "B35":
            scraps[i]["material_number"] = "B22"
        if scraps[i]["material_number"] == "B19-F":
            scraps[i]["material_number"] = "B19"
    
    scraps = unifyMaterialNumbers(scraps, 's')

    raw = [d for d in scraps if d['material_number'].startswith('A')]
    add_scrap(inventory_file, "RAW SCRAP", scrap_date, raw)

    ingredient = [d for d in scraps if d['material_number'].startswith('B')]
    add_scrap(inventory_file, "INGREDIENT SCRAP", scrap_date, ingredient)

    bag = [d for d in scraps if d['material_number'].startswith('C')]
    add_scrap(inventory_file, "BAG SCRAP", scrap_date, bag)

    box = [d for d in scraps if d['material_number'].startswith('D') or d['material_number'].startswith('E')]
    add_scrap(inventory_file, "BOX SCRAP", scrap_date, box)


# def get_latest_consumption_report():
#     # Set the path to the directory containing the consumption reports
#     path = "./consumption reports"
    
#     # Get a list of all xlsx files in the subdirectory
#     xlsx_files = glob.glob(os.path.join(path, "*.xlsx"))
    
#     # Initialize variables to keep track of the latest date and file
#     latest_date = None
#     latest_file = None
    
#     # Loop through each xlsx file
#     for file_path in xlsx_files:
#         # Get just the filename without the path
#         file_name = os.path.basename(file_path)
        
#         # Check if the filename starts with "Consumption Report"
#         if file_name.startswith("Consumption Report"):
#             # Get the date from the filename
#             date_str = file_name[len("Consumption Report")+1:len("Consumption Report")+17]
            
#             # Convert the date string to a datetime object
#             date = datetime.datetime.strptime(date_str, "%Y-%m-%d %H-%M")
            
#             # Check if this file's date is later than the latest date we've seen so far
#             if latest_date is None or date > latest_date:
#                 latest_date = date
#                 latest_file = file_path
    
#     # Return the filename and path of the latest file
#     return latest_file


# def get_latest_scrap_report():
#     """Find the latest Scrap Report file in the 'consumption reports' directory."""
#     folder_path = "./consumption reports"
#     # Get a list of all files in the folder that start with "Scrap Report"
#     files = [f for f in os.listdir(folder_path) if f.startswith("Scrap Report")]
#     # If there are no matching files, return None
#     if not files:
#         return None
#     # Find the file with the latest date in the filename
#     latest_file = None
#     latest_date = None
#     for file in files:
#         try:
#             # Extract the date from the filename and convert it to a datetime object
#             date_str = file[13:23]
#             date = datetime.datetime.strptime(date_str, "%Y-%m-%d")

#             # If this is the first file or the date is later than the previous latest date, update the latest file and date
#             if not latest_date or date > latest_date:
#                 latest_file = file
#                 latest_date = date
#         except:
#             # If there's an error in extracting the date, skip this file
#             continue
#     # Return the full path of the latest file
#     return os.path.join(folder_path, latest_file)


def get_report_date(consumption_report, sheet):
    # Open the given consumption report file
    try:
        wb = openpyxl.load_workbook(consumption_report)
    except:
        printColor('red', f'*** Cannot open {consumption_report}, please check if its current opened ***')

    # Get the WOConsumption worksheet
    ws = wb[sheet]

    # Get the value in cell B4
    date_range = ws["B4"].value

    # Extract the date value between "Date Range: " and " - "
    start_index = date_range.index("Date Range: ") + len("Date Range: ")
    end_index = date_range.index(" - ")
    date_value = date_range[start_index:end_index]

    return date_value


def get_inventory_file(path):
    return './INVENTORY TEMP.xlsx'
    """Search for all files under `path` that start with "INVENTORY", and return the file with the latest modification date."""
    # inventory_files = glob.glob(os.path.join(path, "INVENTORY*.xlsx"))
    # latest_file = None
    # for file in inventory_files:
    #     if not latest_file or os.path.getmtime(file) > os.path.getmtime(latest_file):
    #         latest_file = file
    # return latest_file


# def backup_file(inventory_file):
#     """Create a backup folder with today's date, and copy the inventory file to it with the date as the filename."""
#     # Get today's date in YYYY-MM-DD format
#     backup_folder = datetime.date.today().strftime("%Y-%m-%d")
#     # Create the backup folder if it doesn't already exist
#     backup_path = os.path.join("backups", backup_folder)
#     if not os.path.exists(backup_path):
#         os.makedirs(backup_path)
#     # Construct the filename for the backup file
#     filename = os.path.basename(inventory_file)
#     backup_file = os.path.join(backup_path, filename)
#     # Copy the inventory file to the backup folder
#     try:
#         shutil.copy(inventory_file, backup_file)
#         return True
#     except:
#         return False


def get_consumed_materials(file_path):
    # Open the workbook and select the first worksheet
    try:
        workbook = openpyxl.load_workbook(file_path)
    except:
        printColor('red', f'*** Cannot open {file_path}, please check if its current opened ***')
    worksheet = workbook.worksheets[0]
    
    # Declare an empty list to hold the consumed materials
    consumed_materials = []
    
    # Loop through each row starting from row 9
    for i, row in tqdm(enumerate(worksheet.iter_rows(min_row=9))):
        # Check if the value in column B starts with A, B, C, D, or E
        if row[1].value and re.match(r'W\d', row[1].value):
            break
        if row == worksheet.max_row - 1:
            break
        if row[1].value and row[1].value[0] in ['A', 'B', 'C', 'D', 'E']:
            # print(row[1].value)
            # Get the material number from column B
            material_number = row[1].value.split()[0]
            # print(material_number)
            consumed_material = {'material_number': material_number, 'consumed': find_total(worksheet, i+9)}
            consumed_materials.append(consumed_material)
    
    # Close the workbook
    workbook.close()
    
    # Return the list of consumed materials
    return consumed_materials


def unifyMaterialNumbers(consumption, cat):
    combined = {}
    
    if cat == 'c':
        qty_key = 'consumed'
    elif cat == 's':
        qty_key = 'scraped'

    # Combine elements with the same 'material_number' by adding their 'consumed' values
    for element in consumption:
        material_number = element['material_number']
        qty = element[qty_key]
        if material_number in combined:
            combined[material_number] += qty
        else:
            combined[material_number] = qty

    # Create a new list of dictionaries with the combined results
    output = [{'material_number': material_number, qty_key: qty} for material_number, qty in combined.items()]

    return output


def find_total(worksheet, row_num):
    """Given a worksheet and a row number, find the cell in column C that contains "Totals:" and return the value in column G of the same row."""
    c_col = worksheet['C']
    total_cell = None
    while not total_cell:
        # print(row_num)
        cell = c_col[row_num-1]
        if cell.value == "Totals:":
            total_cell = worksheet.cell(row=row_num, column=7)
        else:
            row_num += 1
    total = total_cell.value.replace(',', '')
    return float(total)


def get_scraped_materials(scrap_report):
    try:
        wb = openpyxl.load_workbook(scrap_report)
    except:
        printColor('red', f'*** Cannot open {scrap_report}, please check if its current opened ***')

    ws = wb["ConsumptionReport"]
    scraped = []
    for row in tqdm(ws.iter_rows(min_row=7)):
        if row[1].value is not None:
            material_number = row[1].value.split()[0]
            scraped_quantity = row[4].value
            scraped.append({"material_number": material_number, "scraped": scraped_quantity})
    return scraped


def get_pulled_materials(consumption_report):
    # Load the workbook
    try:
        wb = openpyxl.load_workbook(consumption_report, read_only=True)
    except:
        printColor('red', f'*** Cannot open {consumption_report}, please check if its current opened ***')

    # Select the WOConsumption worksheet
    ws = wb.worksheets[0]

    current_row = 9

    material_rows = []
    while True:
        cell_value = ws.cell(row=current_row, column=2).value
        print(cell_value)
        sys.stdout.write("\033[F")
        sys.stdout.write("\033[K")
        if cell_value != None and not valid_row(cell_value) and not cell_value.startswith('WO # '):
            break
        elif cell_value != None and valid_row(cell_value):
            material_rows.append(current_row)
        current_row += 1

    pulled_material_groups = []

    for row_number in material_rows:
        pulled_material_groups.append(find_pulled_materials(ws, row_number))

    pulled_materials = []

    for row in pulled_material_groups:
        for element in row:
            pulled_materials.append(element)
    
    for item in pulled_materials:
        if item['lot-num'].endswith('-P'):
            item['lot-num'] = item['lot-num'][:-2]

    pulled_materials = sorted(pulled_materials, key=lambda x: x['lot-num'])
    return pulled_materials


def valid_row(cell_value):
    return cell_value.startswith('A') or cell_value.startswith('B') 
    # or cell_value.startswith('C') or cell_value.startswith('D') or cell_value.startswith('E')


def find_pulled_materials(ws, row_number):
    pulled_material = ws.cell(row=row_number, column=2).value.split()[0]
    pulled_materials = []
    current_row = row_number + 1
    
    while True:
        current_cell = ws.cell(row=current_row, column=2).value
        
        if current_cell is None:
            current_row += 1
            continue
        elif current_cell.startswith('WO #'):
            lot_num = ws.cell(row=current_row, column=2).value
            pattern = r"WO #\s+(.*?):00"
            lot_num = re.search(pattern, lot_num).group(1)
            qty = ws.cell(row=current_row, column=7).value
            pulled_materials.append({'lot-num': lot_num, 'material': pulled_material, 'qty': float(qty.replace(',', ''))})
            printColor('purple', f"{lot_num}: {pulled_material} - {float(qty.replace(',', ''))}")
        elif valid_row(current_cell) or current_cell.startswith('C') or last_row_with_date(current_cell):
            break
        
        current_row += 1
    
    return pulled_materials


def last_row_with_date(current_cell):
    months = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']
    for month in months:
        if current_cell.startswith(month):
            return True
        else:
            return False

def add_pulled_materials(pulled_materials, file, date):
    if len(pulled_materials) == 0: return
    pulled_material_list = [item for item in pulled_materials if item['material'].startswith('A')] + [item for item in pulled_materials if item['material'].startswith('B')]
    
    try:
        wb = openpyxl.load_workbook(file)
    except:
        printColor('red', f'*** Cannot open {file}, please check if its current opened ***')
        
    formatted_date = date.replace('/', '-')
    ws = wb.create_sheet(title=formatted_date)

    for i, material in enumerate(pulled_material_list):
        # printColor('purple', f"{material['lot-num']}: {material['material']} - {material['qty']}")
        row = 1 + i
        # if i == 0 or material['lot-num'] != pulled_material_list[i-1]['lot-num']:
        #     ws.cell(row=row, column=1, value=material['lot-num'])
        if material['material'].startswith('B'):
            row += 1
        ws.cell(row=row, column=1, value=material['lot-num'])
        ws.cell(row=row, column=3, value=material['material'])
        ws.cell(row=row, column=4, value=material['qty'])

    try:
        wb.save(file)
        wb.close()
    except:
        printColor('red', f'***Cannot save {file}, please check if the excel file is currently opened***')


# def add_pulled_materials_to_all(pulled_materials, file):
#     if len(pulled_materials) == 0: return
#     # print(pulled_materials)
#     pulled_material_list = [item for item in pulled_materials if not item['material'].startswith('A')]
#     wb = openpyxl.load_workbook(file)
#     ws = wb['All']

#     last_row = ws.max_row
#     for item in pulled_material_list:
#         ws.cell(row=last_row+1, column=1, value=item['lot-num'])
#         ws.cell(row=last_row+1, column=2, value=item['material'])
#         ws.cell(row=last_row+1, column=3, value=item['qty'])
#         last_row += 1
    
#     # Save changes to the workbook
#     wb.save(file)


def add_scrap(inventory_file, sheet_name, scrap_date, scraps):
    if len(scraps) == 0: return
    col_to_search = 0
    # Open the xlsx file and worksheet
    try:
        wb = openpyxl.load_workbook(inventory_file, read_only=False)
    except:
        printColor('red', f'*** Cannot open {inventory_file}, please check if its current opened ***')
    ws = wb[sheet_name]
    
    # # Search for the column number with value 'scrap' in row 1
    # col_num = None
    # for col in range(1, ws.max_column + 1):
    #     # print(ws.cell(row=1, column=col).value)
    #     if ws.cell(row=1, column=col).value == 'scrap':
    #         col_num = col
    #         break
    
    # if col_num is None:
    #     printColor('red', "Column with value 'scrap' not found in row 1.")
    #     return
    
    # # Loop through each scrap in the scraps list
    # for scrap in scraps:
    #     # Search for the row number with value same as scrap["material_number"] in col_to_search
    #     row_num = None
    #     for row in range(2, ws.max_row + 1):
    #         if ws.cell(row=row, column=col_to_search).value == scrap["material_number"]:
    #             row_num = row
    #             break
        
    #     if row_num is None:
    #         print(f"No row found with material number '{scrap['material_number']}' in column {col_to_search}.")
    #         continue

    #     printColor('purple', f"{scrap['material_number']} - {scrap['scraped']}")
        
    #     # Check if the cell is empty, and set its value accordingly
    #     cell_value = ws.cell(row=row_num, column=col_num).value
    #     scrap_qty = float(scrap['scraped'].replace(',', ''))
    #     if cell_value is None:
    #         ws.cell(row=row_num, column=col_num).value = f"= {scrap_qty}"
    #     else:
    #         ws.cell(row=row_num, column=col_num).value = f"{cell_value} + {scrap_qty}"
    
    # # Save changes to the inventory file
    
    scrap_date = datetime.datetime.strptime(scrap_date, "%m/%d/%Y")
    day_value = int(scrap_date.day)
    
    # Find column number for day_value
    current_column = None
    for col in ws.iter_cols(min_row=1, max_row=1, min_col=1):
        if col[0].value == day_value:
            current_column = col[0].column
            break
    
    for item in scraps:
        # Find row number for material_number
        current_row = None
        for row in ws.iter_rows(min_row=2, max_col=col_to_search + 2):
            # print(row[col_to_search].value, ' - ', item["material_number"])
            excel_material_num = str(row[col_to_search].value).strip()
            if excel_material_num == item["material_number"]:
                if excel_material_num.startswith('A'):
                    color = 'green'
                elif excel_material_num.startswith('B'):
                    color = 'yellow'
                elif excel_material_num.startswith('C') or excel_material_num.startswith('D') or excel_material_num.startswith('E'):
                    color = 'blue'
                printColor(color, f'{excel_material_num} - {item["scraped"]}')
                current_row = row[col_to_search].row
                break
        
        if current_row:
            ws.cell(row=current_row, column=current_column).value = int(item["scraped"])
        
        else:
            printColor('red', f'*** Unable to find {item["material_number"]} in excel file ***')

    wb.save(inventory_file)
    # try:
    #     wb.save(inventory_file)
    
    # except:
    #     printColor('red', f'***Cannot save {inventory_file}, please check if the excel file is currently open***')
    
    wb.close()


def add_consumption(inventory_file, sheet_name, consumpt_date, consumption):
    # print(consumption)
    if len(consumption) == 0: return
    col_to_search = 0
    # Open inventory file with password
    try:
        wb = openpyxl.load_workbook(inventory_file, read_only=False)
    except:
        printColor('red', f'*** Cannot open {inventory_file}, please check if its current opened ***')
    # wb.security = WorkbookProtection(workbookPassword=password)
    ws = wb[sheet_name]

    # Get day value from report_date
    consumpt_date = datetime.datetime.strptime(consumpt_date, "%m/%d/%Y")
    day_value = int(consumpt_date.day)

    # Find column number for day_value
    current_column = None
    for col in ws.iter_cols(min_row=1, max_row=1, min_col=1):
        if col[0].value == day_value:
            current_column = col[0].column
            break

    # print('current_column: ', current_column)

    # Add consumption to inventory
    for item in consumption:
        # Find row number for material_number
        current_row = None
        for row in ws.iter_rows(min_row=2, max_col=col_to_search + 2):
            # print(row[col_to_search].value, ' - ', item["material_number"])
            excel_material_num = str(row[col_to_search].value).strip()
            if excel_material_num == item["material_number"]:
                if excel_material_num.startswith('A'):
                    color = 'green'
                elif excel_material_num.startswith('B'):
                    color = 'yellow'
                elif excel_material_num.startswith('C') or excel_material_num.startswith('D'):
                    color = 'blue'
                printColor(color, f'{excel_material_num} - {item["consumed"]}')
                current_row = row[col_to_search].row
                break
        
        if current_row:
            ws.cell(row=current_row, column=current_column).value = item["consumed"]
        
        else:
            printColor('red', f'*** Unable to find {item["material_number"]} in excel file ***')

    # Save the file
    # wb.security = WorkbookProtection(workbookPassword = '2400LI')
    try:
        wb.save(inventory_file)
        
    except:
        printColor('red', f'***Cannot save {inventory_file}, please check if the excel file is currently open***')
    
    wb.close()


# def add_password(inventory_file, password):
#     return 0


if __name__ == "__main__":
	main()
